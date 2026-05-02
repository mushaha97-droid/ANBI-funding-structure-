"""
Build the cleaned NGO income dataset from raw annual report figures.
Outputs: data/ngo_income_data.xlsx
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Exchange rates (approximate 2023 averages, rounded)
USD_TO_EUR = 0.92
GBP_TO_EUR = 1.15

# Raw data from collected annual reports — one row per (NGO, source_item)
# Categories: donation, business, community_engagement, government, capital
records = [
    # 1. Standaard Publicatieplicht Fondsen (2023)
    ("Standaard Publicatieplicht Fondsen", 2023, "EUR", "Investment Income", "capital", 139774),
    ("Standaard Publicatieplicht Fondsen", 2023, "EUR", "Grants from Other Funds", "government", 75000),
    ("Standaard Publicatieplicht Fondsen", 2023, "EUR", "Contributions from Farms", "community_engagement", 113197),
    ("Standaard Publicatieplicht Fondsen", 2023, "EUR", "Other Income", "capital", 147340),

    # 2. Foundation reNature (2022)
    ("Foundation reNature", 2022, "EUR", "Gifts from small farmers", "community_engagement", 431249),
    ("Foundation reNature", 2022, "EUR", "Interest receivable", "capital", 2568),

    # 3. Stichting Lente Land (2023)
    ("Stichting Lente Land", 2023, "EUR", "Revenue from investments and funds raised", "community_engagement", 286789),
    ("Stichting Lente Land", 2023, "EUR", "Other income", "capital", 13439),

    # 4. Stichting BD Grondbeheer (2023)
    ("Stichting BD Grondbeheer", 2023, "EUR", "Donations (gifts and legacies)", "donation", 322657),
    ("Stichting BD Grondbeheer", 2023, "EUR", "Lease Income from farmers", "community_engagement", 525476),
    ("Stichting BD Grondbeheer", 2023, "EUR", "Perpetual Bonds", "capital", 159500),
    ("Stichting BD Grondbeheer", 2023, "EUR", "Recharged lease (third-party)", "business", 109033),
    ("Stichting BD Grondbeheer", 2023, "EUR", "Rental income", "business", 3600),
    ("Stichting BD Grondbeheer", 2023, "EUR", "Profit on land sales", "business", 2637),
    ("Stichting BD Grondbeheer", 2023, "EUR", "Forgiven loan repayments", "donation", 19588),
    ("Stichting BD Grondbeheer", 2023, "EUR", "Other income (interest, deposits)", "government", 8995),

    # 5. Stichting Herenboeren Nederland (2023)
    ("Stichting Herenboeren Nederland", 2023, "EUR", "Membership Contributions", "community_engagement", 759904),
    ("Stichting Herenboeren Nederland", 2023, "EUR", "One-time Entrance Fees", "business", 300000),
    ("Stichting Herenboeren Nederland", 2023, "EUR", "Grants and Donations", "government", 417000),
    ("Stichting Herenboeren Nederland", 2023, "EUR", "Other Income", "donation", 55000),

    # 6. Stichting Voedselbosbouw Nederland (VBNL) (2023)
    ("Stichting Voedselbosbouw Nederland", 2023, "EUR", "Government and public agency grants", "government", 248022),
    ("Stichting Voedselbosbouw Nederland", 2023, "EUR", "Donations from foundations", "business", 302000),
    ("Stichting Voedselbosbouw Nederland", 2023, "EUR", "Private donations", "donation", 11258),
    ("Stichting Voedselbosbouw Nederland", 2023, "EUR", "Sponsorships and company support", "business", 90000),
    ("Stichting Voedselbosbouw Nederland", 2023, "EUR", "In-kind services and minor projects", "community_engagement", 20020),

    # 7. Coöperatie Land van Ons (2023)
    ("Cooperatie Land van Ons", 2023, "EUR", "Member Contributions (~26,230 members)", "community_engagement", 290000),
    ("Cooperatie Land van Ons", 2023, "EUR", "Land-related Income", "capital", 668000),
    ("Cooperatie Land van Ons", 2023, "EUR", "Government subsidy (Polderlab)", "government", 212000),
    ("Cooperatie Land van Ons", 2023, "EUR", "Grants and donations", "donation", 303000),
    ("Cooperatie Land van Ons", 2023, "EUR", "Interest on savings", "capital", 10000),

    # 8. ROA — Regenerative Organic Alliance (2024) — USD
    ("Regenerative Organic Alliance", 2024, "USD", "Program Services", "community_engagement", 711800),
    ("Regenerative Organic Alliance", 2024, "USD", "Allies Sponsorships", "business", 505000),
    ("Regenerative Organic Alliance", 2024, "USD", "Grants Income", "government", 300000),
    ("Regenerative Organic Alliance", 2024, "USD", "Direct Contributions", "donation", 105000),
    ("Regenerative Organic Alliance", 2024, "USD", "Charitable Contributions", "donation", 38220),
    ("Regenerative Organic Alliance", 2024, "USD", "In-Kind Contributions", "donation", 17220),

    # 9. Urgenda (2023)
    ("Urgenda", 2023, "EUR", "Own fundraising (individuals)", "donation", 1274372),
    ("Urgenda", 2023, "EUR", "Third-party actions / campaigns", "community_engagement", 500000),
    ("Urgenda", 2023, "EUR", "Government funding", "government", 835199),
    ("Urgenda", 2023, "EUR", "Other projects", "business", 139047),
    ("Urgenda", 2023, "EUR", "Non-profit organizations / foundations", "business", 500000),

    # 10. Trees for All (2023)
    ("Trees for All", 2023, "EUR", "Private Individuals", "donation", 1524431),
    ("Trees for All", 2023, "EUR", "Businesses", "business", 6893385),
    ("Trees for All", 2023, "EUR", "Lottery Organisations", "government", 400000),
    ("Trees for All", 2023, "EUR", "Other Non-Profit Organisations", "business", 1331174),

    # 11. Soil Association Group (2023-24) — GBP
    ("Soil Association Group", 2023, "GBP", "Donations & Legacies", "donation", 1958000),
    ("Soil Association Group", 2023, "GBP", "Charitable Activities", "business", 18744000),
    ("Soil Association Group", 2023, "GBP", "Other Trading Activities", "community_engagement", 779000),
    ("Soil Association Group", 2023, "GBP", "Investment Income", "capital", 27000),

    # 12. Commonland (2023)
    ("Commonland", 2023, "EUR", "Particular individuals", "donation", 2000),
    ("Commonland", 2023, "EUR", "Lottery organizations", "government", 1135531),
    ("Commonland", 2023, "EUR", "Other non-profit organizations", "business", 8369684),
    ("Commonland", 2023, "EUR", "Other income", "capital", 10000),

    # 13. Justdiggit Foundation (2023)
    ("Justdiggit Foundation", 2023, "EUR", "Private donors", "donation", 1916300),
    ("Justdiggit Foundation", 2023, "EUR", "Corporate donors", "business", 7364220),
    ("Justdiggit Foundation", 2023, "EUR", "Nonprofit organizations", "business", 3773420),
    ("Justdiggit Foundation", 2023, "EUR", "Governments, grants, subsidies", "government", 608031),
    ("Justdiggit Foundation", 2023, "EUR", "Sale of products or services", "community_engagement", 380000),

    # 14. Natuurmonumenten (Nature miller) (2023)
    ("Natuurmonumenten", 2023, "EUR", "Donations and gifts (individuals)", "donation", 535682),
    ("Natuurmonumenten", 2023, "EUR", "Legacies", "donation", 158153),
    ("Natuurmonumenten", 2023, "EUR", "Companies", "business", 965941),
    ("Natuurmonumenten", 2023, "EUR", "Postcode Loterij", "government", 3564553),
    ("Natuurmonumenten", 2023, "EUR", "Government Subsidies", "government", 729376),
    ("Natuurmonumenten", 2023, "EUR", "Other Nonprofit Organizations", "business", 2373701),
]

# Build long-format dataframe
df = pd.DataFrame(records, columns=[
    "ngo_name", "year", "currency", "source_item", "source_category", "amount_local"
])

# Convert all amounts to EUR for cross-NGO comparison
def to_eur(row):
    if row["currency"] == "EUR":
        return row["amount_local"]
    elif row["currency"] == "USD":
        return row["amount_local"] * USD_TO_EUR
    elif row["currency"] == "GBP":
        return row["amount_local"] * GBP_TO_EUR
    return None

df["amount_eur"] = df.apply(to_eur, axis=1).round(0).astype(int)

# Sort
df = df.sort_values(["ngo_name", "source_category"]).reset_index(drop=True)

# ------------- Build the Excel workbook -------------
wb = Workbook()

# === Sheet 1: README ===
ws = wb.active
ws.title = "README"
header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="2F5496")
section_font = Font(name="Arial", size=12, bold=True, color="2F5496")
body_font = Font(name="Arial", size=11)

ws["A1"] = "NGO Income Source Dataset — RAAP! Fundraising Benchmark Study"
ws["A1"].font = header_font
ws["A1"].fill = header_fill
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28
ws.merge_cells("A1:B1")

readme_rows = [
    ("", ""),
    ("Purpose", "Benchmark dataset of income sources from NGOs working in regenerative agriculture, nature restoration, and sustainable land use. Used to inform RAAP!'s fundraising strategy."),
    ("", ""),
    ("Sheets", ""),
    ("Raw Data", "One row per (NGO, income source item). Long format. The unit of analysis."),
    ("Summary by NGO", "Pivoted view: one row per NGO with totals and category breakdown."),
    ("Categories", "Definition of the five income source categories used in this analysis."),
    ("", ""),
    ("Categories used", "donation | business | community_engagement | government | capital"),
    ("", ""),
    ("Currency handling", "All amounts converted to EUR for comparison."),
    ("USD to EUR rate", USD_TO_EUR),
    ("GBP to EUR rate", GBP_TO_EUR),
    ("Note", "2023 average rates. Original local-currency values preserved in the 'currency' and 'amount_local' columns."),
    ("", ""),
    ("Data sources", "Public annual reports and Standaard Publicatieplicht filings (2022-2024)."),
    ("NGOs included", df["ngo_name"].nunique()),
    ("Source records", len(df)),
    ("Reporting years covered", "2022 - 2024"),
]

row = 2
for label, val in readme_rows:
    ws.cell(row=row, column=1, value=label).font = section_font if val == "" and label else body_font
    if label and val == "":
        ws.cell(row=row, column=1).font = section_font
    ws.cell(row=row, column=2, value=val).font = body_font
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 90

# === Sheet 2: Raw Data ===
ws2 = wb.create_sheet("Raw Data")
headers = ["ngo_name", "year", "currency", "source_item", "source_category",
           "amount_local", "amount_eur"]
for col, h in enumerate(headers, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="2F5496")
    c.alignment = Alignment(horizontal="center")

for r_idx, row_data in enumerate(df.itertuples(index=False), start=2):
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name="Arial", size=11)
        if c_idx in (6, 7):  # amount cols
            cell.number_format = '#,##0;(#,##0);-'

# Freeze header & widths
ws2.freeze_panes = "A2"
widths = [34, 8, 10, 50, 24, 16, 16]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# === Sheet 3: Summary by NGO ===
ws3 = wb.create_sheet("Summary by NGO")
pivot = df.pivot_table(
    index="ngo_name", columns="source_category", values="amount_eur",
    aggfunc="sum", fill_value=0
).reset_index()

# Ensure all 5 category columns exist
for cat in ["donation", "business", "community_engagement", "government", "capital"]:
    if cat not in pivot.columns:
        pivot[cat] = 0

# Reorder
pivot = pivot[["ngo_name", "donation", "business", "community_engagement", "government", "capital"]]

# Year for each NGO
year_map = df.groupby("ngo_name")["year"].first().to_dict()
pivot.insert(1, "year", pivot["ngo_name"].map(year_map))

summary_headers = list(pivot.columns) + ["total_eur"]
for col, h in enumerate(summary_headers, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="2F5496")
    c.alignment = Alignment(horizontal="center")

for r_idx, row_data in enumerate(pivot.itertuples(index=False), start=2):
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name="Arial", size=11)
        if c_idx >= 3:
            cell.number_format = '#,##0;(#,##0);-'
    # total formula in last column
    total_cell = ws3.cell(row=r_idx, column=len(summary_headers))
    total_cell.value = f"=SUM(C{r_idx}:G{r_idx})"
    total_cell.number_format = '#,##0;(#,##0);-'
    total_cell.font = Font(name="Arial", size=11, bold=True)

ws3.freeze_panes = "A2"
ws3.column_dimensions["A"].width = 34
ws3.column_dimensions["B"].width = 8
for col in ["C", "D", "E", "F", "G", "H"]:
    ws3.column_dimensions[col].width = 18

# === Sheet 4: Categories ===
ws4 = wb.create_sheet("Categories")
cat_data = [
    ("Category", "Definition", "Examples"),
    ("donation", "Voluntary unrestricted gifts from individuals or estates",
     "Private donor gifts, legacies, one-off contributions"),
    ("business", "Income from corporate sponsors, foundations, or commercial activity",
     "Corporate sponsorships, foundation grants, sales of services to companies"),
    ("community_engagement", "Income tied to community participation, membership, or producer relationships",
     "Member fees, lease payments from farmers, program services, in-kind contributions"),
    ("government", "Public funding from national/EU government, agencies, or lottery bodies",
     "EU grants, ministry subsidies, RVO, Postcode Loterij (statutory)"),
    ("capital", "Returns on assets, investments, or land",
     "Investment income, interest on savings, perpetual bonds, land returns"),
]
for r_idx, row_vals in enumerate(cat_data, 1):
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws4.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="2F5496")
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.font = Font(name="Arial", size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 50
ws4.column_dimensions["C"].width = 55

# Save
out_path = "data/ngo_income_data.xlsx"
wb.save(out_path)
print(f"Saved {out_path}")
print(f"NGOs: {df['ngo_name'].nunique()}, records: {len(df)}")
